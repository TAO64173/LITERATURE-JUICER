"""Excel 写入模块

将论文提取结果写入结构化 Excel 文件。
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 表头定义（21 列）
HEADERS = [
    "Title", "Authors", "Year", "Journal", "DOI", "Keywords", "Abstract",
    "研究问题", "研究背景", "研究动机", "研究目标",
    "研究方法", "数据集/实验设置", "性能指标", "对比方法",
    "创新点", "主要发现", "结论",
    "局限性", "未来工作", "可借鉴点/启发",
]

HEADER_KEYS = [
    "title", "author", "year", "journal", "doi", "keywords", "abstract",
    "question", "background", "gap", "objective",
    "method", "dataset", "metrics", "comparison",
    "innovation", "findings", "conclusion",
    "limitation", "future_work", "inspiration",
]

# 样式
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COL_WIDTHS = [
    30, 18, 10, 18, 18, 25, 35,   # Title, Authors, Year, Journal, DOI, Keywords, Abstract
    35, 30, 30, 30,                # 研究问题, 研究背景, 研究动机, 研究目标
    35, 30, 30, 30,                # 研究方法, 数据集, 性能指标, 对比方法
    30, 30, 30,                    # 创新点, 主要发现, 结论
    30, 30, 30,                    # 局限性, 未来工作, 可借鉴点
]


def write_excel(papers: list[dict], output_path: str) -> str:
    """
    将论文数据写入 Excel 文件

    Args:
        papers: 论文数据列表，每项包含 HEADER_KEYS 中的所有字段
        output_path: 输出文件路径

    Returns:
        输出文件的绝对路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "文献矩阵"

    # 写表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 设置列宽
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 写数据
    for row_idx, paper in enumerate(papers, 2):
        for col_idx, key in enumerate(HEADER_KEYS, 1):
            value = paper.get(key, "未明确提及")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    # 冻结首行
    ws.freeze_panes = "A2"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out.resolve())
